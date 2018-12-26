__author__ = 'lulu614'

from openpyxl import load_workbook
from bin.Class import *
from bin.find_file import find_file



'''
    导入全部参数
'''

def load_pile_dict():
    '''
        导入所有桩参数
        :return: 所有桩参数
    '''

    #打开Excel表格
    wb = load_workbook('pile.xlsx')
    ws = wb.active

    pile_dict = Pile_dict()
    for i in ws.iter_rows(min_row=2):
        if i[0].value != None:
            num = i[0].value
            x = i[1].value/1000
            y = i[2].value/1000
            Q = i[3].value
            l = i[4].value
            d = i[5].value
            E = i[6].value
            pile = Pile(num,x,y,Q,l,d,E)
            pile_dict.add(pile)
        else:break
    wb.close()
    return pile_dict


def load_soil_list():
    '''
        导入所有土层参数
        :return: 所有土层参数
    '''

    wb = load_workbook('soil.xlsx')
    ws = wb.active
    soil_list = Soil_list()

    for i in ws.iter_rows(min_row=2):
        if i[0].value != None:
            num = i[0].value
            name = i[1].value
            h = i[2].value
            Es = i[3].value
            v = i[4].value
            soil = Soil(num,name,h,Es,v)
            soil_list.add(soil)
        else:break
    wb.close()
    return soil_list

def load_para():
    '''
        导入其他参数
        :return: 其他参数
                    empirical_coefficient = 经验系数
                    alph = 桩端阻力比
                    bata = 沿桩身均匀分布的桩侧摩阻力比
                    sphere_influence = 群桩影响范围
                    layer_thickness = 土层分层厚度
                    depth = 桩顶深度
    '''
    wb = load_workbook('para.xlsx')
    ws = wb.active

    para_list = [1,0.2,0,20,2,5.5]
    for i in range(6):
        para_list[i] = ws.cell(row=i+1,column=3).value
    [empirical_coefficient, alph, bata, sphere_influence, layer_thickness,depth] = para_list
    para = Para(empirical_coefficient, alph, bata, sphere_influence, layer_thickness,depth)
    wb.close()
    return para

if __name__ == '__main__':
    find_file()
    pile_dict= load_pile_dict()
    # print(pile_dict.pile_dict)
    soil_list = load_soil_list()
    print(soil_list.soil_list[9].v)
    para = load_para()
    print(para)